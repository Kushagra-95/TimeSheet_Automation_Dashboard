from pathlib import Path
import sys
import pandas as pd
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import zipfile
import shutil
import tempfile
import xml.etree.ElementTree as ET
# -----------------------------
# Read paths from command line
# -----------------------------
input_dir = Path(sys.argv[1])
output_dir = Path(sys.argv[2])
all_timesheets_folder = output_dir / "All Timesheets"
all_timesheets_folder.mkdir(parents=True, exist_ok=True)
# Find uploaded Excel file
excel_files = list(input_dir.glob("*.xlsx"))

if not excel_files:
    raise FileNotFoundError(f"No Excel file found in {input_dir}")

input_file = excel_files[0]

def repair_excel(input_file):
    """
    Repairs SAP generated xlsx files that contain invalid
    empty <fill/> elements inside styles.xml.
    Returns the repaired workbook path.
    """

    temp_dir = Path(tempfile.mkdtemp())

    with zipfile.ZipFile(input_file, "r") as z:
        z.extractall(temp_dir)

    styles = temp_dir / "xl" / "styles.xml"

    if styles.exists():

        ns = {
            "x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        }

        tree = ET.parse(styles)
        root = tree.getroot()

        fills = root.find("x:fills", ns)

        if fills is not None:

            changed = False

            for fill in fills.findall("x:fill", ns):

                if len(fill) == 0:

                    pattern = ET.SubElement(
                        fill,
                        "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}patternFill"
                    )

                    pattern.set("patternType", "none")

                    changed = True

            if changed:
                tree.write(
                    styles,
                    encoding="utf-8",
                    xml_declaration=True
                )

    repaired = temp_dir / "repaired.xlsx"

    with zipfile.ZipFile(
        repaired,
        "w",
        zipfile.ZIP_DEFLATED
    ) as new_zip:

        for file in temp_dir.rglob("*"):

            if file.is_file() and file != repaired:

                new_zip.write(
                    file,
                    file.relative_to(temp_dir)
                )

    return repaired

def format_name(email):
    try:
        local_part = email.split("@")[0]
        first, last = local_part.split(".")
        first = "".join(filter(str.isalpha, first))
        last = "".join(filter(str.isalpha, last))
        return f"{first.title()} {last.title()}"
    except Exception:
        return email


# Read second sheet
try:

    df = pd.read_excel(
        input_file,
        sheet_name=1,
        engine="openpyxl",
        dtype=str
    )

except Exception as e:

    print("Repairing workbook...")

    repaired = repair_excel(input_file)

    df = pd.read_excel(
        repaired,
        sheet_name=1,
        engine="openpyxl",
        dtype=str
    )

# Select required columns
selected_columns = df.iloc[:, [8, 10, 13, 20]].copy()

selected_columns.columns = [
    "Name",
    "Timesheet Date",
    "Recorded Hours",
    "Note"
]

selected_columns["Name"] = (
    selected_columns["Name"]
    .apply(format_name)
    .str.strip()
    .str.title()
)

# Split by Name
for name in selected_columns["Name"].unique():

    filtered_df = selected_columns[
        selected_columns["Name"] == name
    ]

    # safe_name = "".join(filter(str.isalpha, name))

    # output_path = output_dir / f"{safe_name}.xlsx"
    month_year=pd.to_datetime(
        filtered_df["Timesheet Date"].iloc[0]
    ).strftime("%B %Y")

    safe_name="".join(c for c in name if c.isalnum() or c == " ").strip()
    file_name=f"{month_year} Timesheet {safe_name}.xlsx"
    output_path=output_dir/file_name
    folder_output_path = all_timesheets_folder / file_name
    for save_path in [output_path, folder_output_path]:

        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:

            filtered_df.to_excel(
                writer,
                index=False,
                sheet_name="Sheet1"
            )

            worksheet = writer.sheets["Sheet1"]

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            for col_idx, column in enumerate(filtered_df.columns, 1):

                col_letter = get_column_letter(col_idx)

                max_length = max(
                    filtered_df[column].astype(str).map(len).max(),
                    len(column)
                ) + 2

                worksheet.column_dimensions[col_letter].width = max_length

                for row in range(2, len(filtered_df) + 2):
                    worksheet[f"{col_letter}{row}"].border = thin_border

            for col_idx in range(1, len(filtered_df.columns) + 1):
                worksheet[f"{get_column_letter(col_idx)}1"].border = thin_border

print(f"{len(selected_columns['Name'].unique())} individual timesheets generated successfully.")